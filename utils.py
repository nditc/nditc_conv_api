from io import BytesIO
import json
import threading
import time
from openpyxl import Workbook, load_workbook
import os, uuid
from weasyprint import HTML


token_map = {}
TEMP_DIR = "./files"
os.makedirs(TEMP_DIR, exist_ok=True)


def schedule_cleanup(token: str, delay: int = 60):
    def cleanup():
        time.sleep(delay)
        path = token_map.pop(token, None)
        if path and os.path.exists(path):
            os.remove(path)
    threading.Thread(target=cleanup, daemon=True).start()

def json_to_xlsx(json_data, method_id):
    token = str(uuid.uuid4())
    filepath = os.path.join(TEMP_DIR, f"{token}.xlsx")

    all_keys = set()
    for item in json_data:
        all_keys.update(item.keys())
    headers = list(all_keys)

    wb = Workbook()
    ws = wb.active

    ws.append(headers)

    for item in json_data:
        row = []
        for key in headers:
            value = item.get(key, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            row.append(value)
        ws.append(row)

    if method_id == 1:
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream
    else:
        token_map[token] = filepath
        schedule_cleanup(token)
        wb.save(filepath)
        return token


def xlsx_to_json(xlsx_datastream):
    wb = load_workbook(BytesIO(xlsx_datastream))
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] for i in range(len(headers))}
        data.append(item)
    return data

def html_to_pdf(html_str, method_id):
    token = str(uuid.uuid4())
    pdfpath = os.path.join(TEMP_DIR, f"{token}.pdf")
    print(pdfpath)
    pdf = HTML(string=html_str).write_pdf()

    if method_id == 1:
        pdf_stream = BytesIO(pdf)
        return pdf_stream
    else:
        with open(pdfpath, "wb") as f:
            f.write(pdf)
        token_map[token] = pdfpath
        schedule_cleanup(token)
        return token

if __name__ == "__main__":
    json_to_xlsx(json_test_data)
    # data = xlsx_to_json("output.xlsx")
    # print(data)